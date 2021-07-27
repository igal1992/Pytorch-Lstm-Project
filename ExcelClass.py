import openpyxl
from pathlib import Path
import numpy as np


# init openpyxl, and return active sheet
def init_openpyxl(nameofFile):
    xlsx_file = Path(nameofFile + '.xlsx')
    wb_obj = openpyxl.load_workbook(xlsx_file)
    return wb_obj.active


class ExcelClass:

    def __init__(self, rowIndexOfTitles, nameofFile):
        self.sheet = init_openpyxl(nameofFile)  # Read the active sheet:
        self.rowIndexOfTitles = rowIndexOfTitles
        self.counter = 0
        self.startingIndex = 3
        self.titleIndex = {
            "time": 1,
            "azimuth": 2,
            "elevation": 3,
            "range": 4,
            "EIRP": 5,
            "XmtrPower": 6,
            "XmtrGain": 7,
            "XmtrEIRP": 8,
            "atmosphereLoss": 9,
            "urbanTerresLoss": 10,
            "urbanTerresLossPred": 11,
            "rainLoss": 12,
            "cloudsFogLoss": 13,
            "tropoScintillLose": 14,
            "ionoFadingLoss": 15,
            "freq_dopplerShift": 16,
            "tearth": 17,
            "propagationDelay": 18,
            "propagationDistance": 19,
            "rcvd_iso_power": 20,
            "c_n": 21,
            "c_no": 22,
            "carrierPowerAtRcvrInput": 23,
            "rcvrGain": 24,
            "eb_no": 25,
            "tatmos": 26,
            "t_urbanTerres": 27,
            "train": 28,
            "t_cloudsFog": 29,
            "t_tropoScintill": 30,
            "t_ionoFading": 31
        }
        self.data = {
            "time": [],
            "azimuth": [],
            "elevation": [],
            "range": [],
            "EIRP": [],
            "XmtrPower": [],
            "XmtrGain": [],
            "XmtrEIRP": [],
            "atmosphereLoss": [],
            "urbanTerresLoss": [],
            "rainLoss": [],
            "cloudsFogLoss": [],
            "tropoScintillLose": [],
            "ionoFadingLoss": [],
            "freq_dopplerShift": [],
            "tearth": [],
            "propagationDelay": [],
            "propagationDistance": [],
            "rcvd_iso_power": [],
            "c_n": [],
            "c_no": [],
            "carrierPowerAtRcvrInput": [],
            "rcvrGain": [],
            "eb_no": [],
            "tatmos": [],
            "t_urbanTerres": [],
            "train": [],
            "t_cloudsFog": [],
            "t_tropoScintill": [],
            "t_ionoFading": []
        }

    # return an array that contains all the column's values. only for a single session
    def getDataFromColumn(self, columnName, starting_row):
        arrToReturn = []
        columnIndex = self.getIndexOfColumn(columnName)
        for i in range(starting_row, self.sheet.max_row + 1):
            cellValue = self.sheet.cell(i, columnIndex)
            if cellValue.value is None:
                return arrToReturn
            arrToReturn.append(cellValue.value)
        print("No NONE value has found")
        return arrToReturn

    # update one column in the excel with given array in a single session
    def setDataInExcel(self, columnName, array, fileName,preds):
        xlsx_file = Path(fileName+'.xlsx')  # create a ref path
        wb_obj = openpyxl.load_workbook(xlsx_file)  # load the excel with the ref path
        if preds:
            columnIndex = self.getIndexOfColumn(columnName) + 1  # get index of column by name
        else:
            columnIndex = self.getIndexOfColumn(columnName)  # get index of column by name
        for i in range(self.startingIndex, wb_obj.active.max_row + 1):
            cellValue = wb_obj.active.cell(i, columnIndex)
            if cellValue.value is None:
                print("END")
                wb_obj.save(fileName+'.xlsx')
                return
            if preds:
                cellValue = wb_obj.active.cell(i, columnIndex-1)
                cellValue.value = array[i - self.startingIndex]
            else:
                cellValue.value = array[i - self.startingIndex]
        print("No NONE value has found")

    # stores given array data in the right array(Column) in the dictionary
    def storeArrayColumnInArray(self, arrToStore, columnName):
        self.data[columnName] = arrToStore

    # stores column data in the right array in the dictionary
    def storeColumnInArray(self, columnName):
        self.data[columnName] = self.getDataFromColumn(columnName, self.startingIndex)

    # stores all the data of a day in dictionary without filter.
    def storeAllDataByDefault(self):
        return

    # return the index of the key in the dictionary
    def getIndexOfColumn(self, columnName):
        return self.titleIndex[columnName]

    # accoriding to the column and the session sets the starting Row index of the class
    def setStartingRow(self, columnName, session_num):
        columnIndex = self.getIndexOfColumn(columnName)
        for i in range(self.startingIndex, self.sheet.max_row + 1):
            cellValue = self.sheet.cell(i, columnIndex)
            if isinstance(cellValue.value, float) or isinstance(cellValue.value, int):
                if self.counter == session_num:
                    self.startingIndex = i
                    return i
            elif cellValue.value is None or isinstance(cellValue.value, str):
                temp = self.sheet.cell(i + 1, columnIndex)
                if isinstance(temp.value, float) or isinstance(temp.value, int):
                    self.counter += 1

    # the parameters are vectors (arrays)
    def get_data_matrix(self, index):
        azimuth = self.getDataFromColumn("azimuth", index)
        elevation = self.elevationFix("data72")
        _range = self.getDataFromColumn("range", index)
        eirp = self.getDataFromColumn("EIRP", index)
        xmtr_gain = self.getDataFromColumn("XmtrGain", index)
        xmtr_gain_intensity = self.getDataFromColumn("XmtrEIRP", index)
        urban_terres_loss = self.getDataFromColumn("urbanTerresLoss", index)
        input_matrix = [azimuth, elevation, _range, eirp, xmtr_gain, xmtr_gain_intensity]
        input_matrix = np.transpose(input_matrix)

        final_matrix = []
        for i in range(len(input_matrix)):
            final_matrix.append([input_matrix[i], urban_terres_loss[i]])

        return final_matrix


    # fixes the elevation of the excel according to the input
    def elevationFix(self,nameofFile):
        exlObject = ExcelClass(2, nameofFile)
        index = self.startingIndex
        arr = exlObject.getDataFromColumn("elevation", index)
        mountainTop = 0
        for i in range(len(arr)):
            if arr[i] > arr[i + 1] and i + 1 < len(arr):
                mountainTop = i
                break
        for i in range(mountainTop + 1, len(arr)):
            arr[i] = 180 - arr[i]
        return arr


if __name__ == '__main__':
    print("TEST EXCEL MODULE...")
    exlObject = ExcelClass(2)
    exlObject.startingIndex = 3
    temp = exlObject.getDataFromColumn("urbanTerresLoss", 3)
    array = [1] * len(temp)
    exlObject.setDataInExcel("urbanTerresLoss", array)
